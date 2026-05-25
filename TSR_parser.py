#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import os
import zipfile
import io
from copy import copy
from pathlib import Path
from typing import List, Tuple, Optional, Dict

from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import logging

# ----------------------------------------------------------------------
# Настройка логгера (цветные метки в консоли)
# ----------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)


class _ColorFormatter(logging.Formatter):
    """Простейший цветовой форматтер – только для красоты."""
    COLORS = {
        "INFO": "\033[96m",      # cyan
        "WARNING": "\033[93m",   # yellow
        "ERROR": "\033[91m",     # red
        "DEBUG": "\033[92m",     # green
    }
    RESET = "\033[0m"

    def format(self, record):
        level = record.levelname
        if level in self.COLORS:
            record.levelname = f"{self.COLORS[level]}{level}{self.RESET}"
        return super().format(record)


for h in logging.root.handlers:
    h.setFormatter(_ColorFormatter("%(asctime)s %(levelname)s %(message)s"))
LOGGER = logging.getLogger(__name__)

# ----------------------------------------------------------------------
# Константы
# ----------------------------------------------------------------------
TARGET_NAME = "sysinfo_DCIM_View.xml" ## 
OUT_SUBDIR = "extracted_xml"

# ----------------------------------------------------------------------
# Вспомогательные функции
# ----------------------------------------------------------------------
def ask_path(prompt_text: str) -> Path:
    """Запрос пути у пользователя (проверка, что строка не пустая)."""
    while True:
        raw = input(prompt_text).strip()
        if raw:
            return Path(raw).expanduser().resolve()
        print("⚠️  Путь не может быть пустым – попробуйте ещё раз.")


def ask_path_or_arg() -> Path:
    """Получаем каталог с архивами: из аргументов CLI или через ввод."""
    if len(sys.argv) > 1:
        return Path(sys.argv[1]).expanduser().resolve()
    return ask_path("Введите путь к каталогу с zip‑файлами: ")


# ----------------------------------------------------------------------
# Поиск XML‑файлов внутри (вложенных) zip‑архивов
# ----------------------------------------------------------------------
def find_in_zip(
    zf: zipfile.ZipFile,
    target: str = TARGET_NAME,
    path_prefix: str = "",
) -> List[Tuple[bytes, str]]:
    """
    Рекурсивно ищет `target` во всех вложенных zip‑файлах.
    Возвращает список кортежей (bytes‑данные, «виртуальный путь»).
    """
    results: List[Tuple[bytes, str]] = []

    for entry in zf.infolist():
        entry_path = f"{path_prefix}{entry.filename}"

        # найденный файл‑цель
        if os.path.basename(entry.filename) == target and not entry.is_dir():
            with zf.open(entry, "r") as f:
                results.append((f.read(), entry_path))

        # вложенный zip‑архив
        elif entry.filename.lower().endswith(".zip") and not entry.is_dir():
            with zf.open(entry, "r") as nested_f:
                nested_bytes = nested_f.read()
            safe_path = entry_path.replace("/", "_").replace("\\", "_")
            try:
                with zipfile.ZipFile(io.BytesIO(nested_bytes)) as nested_zf:
                    results.extend(find_in_zip(nested_zf, target, entry_path + "__"))
            except zipfile.BadZipFile:
                LOGGER.warning("Bad nested zip: %s", entry_path)

    return results


def extract_from_archive(archive_path: Path, out_dir: Path, target: str = TARGET_NAME) -> None:
    """Разархивирует все найденные `target` из `archive_path`."""
    LOGGER.info("Processing archive: %s", archive_path)

    try:
        with zipfile.ZipFile(archive_path, "r") as zf:
            found = find_in_zip(zf, target=target)
    except zipfile.BadZipFile:
        LOGGER.error("Cannot open %s as zip.", archive_path)
        return

    if not found:
        LOGGER.info("  → %s not found", target)
        return

    out_dir.mkdir(parents=True, exist_ok=True)

    for data, virtual_path in found:
        safe_path = virtual_path.replace("/", "_").replace("\\", "_").replace("::", "__")
        out_name = f"{archive_path.stem}__{safe_path}"
        out_file = out_dir / out_name
        LOGGER.info("  → extracting to %s", out_file)
        out_file.write_bytes(data)


# ----------------------------------------------------------------------
# Парсинг XML‑файла
# ----------------------------------------------------------------------
def _get_value(root: etree._Element, prop_name: str) -> Optional[str]:
    node = root.find(f".//PROPERTY[@NAME='{prop_name}']/VALUE")
    return node.text.strip() if node is not None and node.text else None


def _get_model_from_system_view(root: etree._Element) -> Optional[str]:
    """Модель берётся из DCIM_SystemView → Property 'Model'."""
    sv = root.find(".//INSTANCE[@CLASSNAME='DCIM_SystemView']")
    if sv is None:
        return None
    node = sv.find(".//PROPERTY[@NAME='Model']/VALUE")
    return node.text.strip() if node is not None and node.text else None


def _get_sn_from_system_view(root: etree._Element) -> Optional[str]:
    """
    Серийный номер – значение свойства `ChassisServiceTag`
    внутри DCIM_SystemView.
    """
    sv = root.find(".//INSTANCE[@CLASSNAME='DCIM_SystemView']")
    if sv is None:
        return None
    node = sv.find(".//PROPERTY[@NAME='ChassisServiceTag']/VALUE")
    if node is not None and node.text:
        return node.text.strip()
    # Фолбэк – старое поле (на всякий случай)
    return _get_value(root, "SerialNumber")


def _extract_disk_part_number(inst: etree._Element) -> str:
    """
    Извлекает парт-номер диска из поля PPID по новым правилам.
    Если PPID нет или он некорректный, использует старый способ.
    """
    # Пытаемся найти PPID
    ppid_node = inst.find(".//PROPERTY[@NAME='PPID']/VALUE")
    if ppid_node is not None and ppid_node.text:
        ppid = ppid_node.text.strip()

        # Удаляем все тире и пробелы
        cleaned_ppid = ppid.replace("-", "").replace(" ", "")

        # Извлекаем 6 символов, начиная с 3-го
        if len(cleaned_ppid) >= 9:  # Нужно как минимум 3 + 6 = 9 символов
            part_number = cleaned_ppid[2:8]  # Символы с 3 по 8 (6 символов)
            return part_number

    # Если PPID не найден или некорректен, используем старый способ
    pn_node = inst.find(".//PROPERTY[@NAME='PartNumber']/VALUE")
    if pn_node is not None and pn_node.text:
        raw_pn = pn_node.text.strip()
        return _format_part_number(raw_pn)

    return ""


# ----------------------------------------------------------------------
# Коллекторы компонентов
# ----------------------------------------------------------------------
def _collect_memory(root: etree._Element) -> List[Tuple[str, str, int]]:
    """
    Считывает каждую планку памяти и формирует строку:
    "<size>GB <Model> <Speed> <Manufacturer>"
    Spare – значение свойства PartNumber.
    """
    counter: Dict[Tuple[str, str], int] = {}

    for inst in root.findall(".//INSTANCE[@CLASSNAME='DCIM_MemoryView']"):
        # ----- Size (в MB → GB) -----
        size_node = inst.find(".//PROPERTY[@NAME='Size']/VALUE")
        size_gb = "0GB"
        if size_node is not None and size_node.text:
            try:
                num = int(size_node.text.split()[0])          # например "16384"
                size_gb = f"{round(num / 1024)}GB"            # 16384 MB → 16 GB
            except Exception:
                pass

        # ----- Model (тип памяти) -----
        model_node = inst.find(".//PROPERTY[@NAME='Model']/VALUE")
        model = model_node.text.strip() if model_node is not None and model_node.text else ""

        # ----- Manufacturer (производитель) -----
        man_node = inst.find(".//PROPERTY[@NAME='Manufacturer']/VALUE")
        manufacturer = man_node.text.strip() if man_node is not None and man_node.text else ""

        # ----- Speed (частота) -----
        speed_node = inst.find(".//PROPERTY[@NAME='Speed']/VALUE")
        speed = speed_node.text.strip() if speed_node is not None and speed_node.text else ""

        # ----- PartNumber (Spare) -----
        pn_node = inst.find(".//PROPERTY[@NAME='PartNumber']/VALUE")
        part_number = pn_node.text.strip() if pn_node is not None and pn_node.text else ""

        # Формируем описание
        description_parts = [size_gb, model, speed, manufacturer]
        description = " ".join(filter(None, description_parts))

        key = (description, part_number)
        counter[key] = counter.get(key, 0) + 1

    # Возврат в формате (Description, Spare, Quantity)
    return [(descr, pn, qty) for (descr, pn), qty in sorted(counter.items())]


def _format_part_number(raw_pn: str) -> str:
    """
    Форматирует парт-номер: удаляет первый и последние 3 символа.
    Примеры:
    0FR0KXA02 -> FR0KX
    0TC67CA01 -> TC67C
    ABCDEFGHI -> BCDEF (если длина >= 5)
    """
    if not raw_pn:
        return ""

    # Удаляем первый символ
    if len(raw_pn) > 1:
        formatted = raw_pn[1:]
    else:
        return raw_pn

    # Удаляем последние 3 символа, если длина позволяет
    if len(formatted) > 3:
        formatted = formatted[:-3]

    return formatted


def _get_motherboard_part_number(root: etree._Element) -> Optional[str]:
    """
    Извлекает и форматирует парт-номер материнской платы из DCIM_SystemView.
    Удаляет первый и последние 3 символа.
    """
    sv = root.find(".//INSTANCE[@CLASSNAME='DCIM_SystemView']")
    if sv is None:
        return None

    # Ищем поле BoardPartNumber
    pn_node = sv.find(".//PROPERTY[@NAME='BoardPartNumber']/VALUE")
    if pn_node is not None and pn_node.text:
        raw_part_number = pn_node.text.strip()
        return _format_part_number(raw_part_number)

    return None


def _collect_cpus(root: etree._Element) -> List[Tuple[str, str, int]]:
    counter: Dict[str, int] = {}
    for inst in root.findall(".//INSTANCE[@CLASSNAME='DCIM_CPUView']"):
        model_node = inst.find(".//PROPERTY[@NAME='Model']/VALUE")
        if model_node is not None and model_node.text:
            model = model_node.text.strip()
            counter[model] = counter.get(model, 0) + 1
    return [("CPU", model, qty) for model, qty in sorted(counter.items())]


def _collect_psus(root: etree._Element) -> List[Tuple[str, str, int]]:
    """
    Собирает информацию о блоках питания с форматированием парт-номеров.
    Удаляет первый и последние 3 символа из парт-номеров.
    """
    counter: Dict[str, int] = {}
    for inst in root.findall(".//INSTANCE[@CLASSNAME='DCIM_PowerSupplyView']"):
        pn_node = inst.find(".//PROPERTY[@NAME='PartNumber']/VALUE")
        if pn_node is not None and pn_node.text:
            raw_pn = pn_node.text.strip()
            formatted_pn = _format_part_number(raw_pn)
            counter[formatted_pn] = counter.get(formatted_pn, 0) + 1

    return [("PSU", pn, qty) for pn, qty in sorted(counter.items())]


def _bytes_to_gb(size_bytes_str: str) -> str:
    """«123456 Bytes» → «123 GB» (округляем)."""
    try:
        number_part = size_bytes_str.split()[0]
        size_int = int(number_part)
        gb = round(size_int / (1024 ** 3))
        return f"{gb}GB"
    except Exception:
        return "0GB"


def _collect_disks(root: etree._Element) -> List[Tuple[str, str, int]]:
    """
    (Description, Spare, Quantity) для дисков.
    Теперь парт-номер извлекается из PPID по новым правилам.
    """
    counter: Dict[Tuple[str, str], int] = {}

    for inst in root.findall(".//INSTANCE[@CLASSNAME='DCIM_PhysicalDiskView']"):
        # Spare – извлекаем парт-номер из PPID или PartNumber
        part_number = _extract_disk_part_number(inst)

        # Model (для отладки или резервного варианта)
        model_node = inst.find(".//PROPERTY[@NAME='Model']/VALUE")
        model = model_node.text.strip() if model_node is not None and model_node.text else ""

        # Size
        size_node = inst.find(".//PROPERTY[@NAME='SizeInBytes']/VALUE")
        size_str = size_node.text.strip() if size_node is not None and size_node.text else "0 Bytes"
        size_gb = _bytes_to_gb(size_str)

        # MediaType (1 → SSD)
        media_node = inst.find(".//PROPERTY[@NAME='MediaType']")
        if media_node is not None:
            media_value = media_node.find("VALUE").text if media_node.find("VALUE") is not None else ""
            if media_value == "1":
                media = "SSD"
            else:
                media = media_node.find("DisplayValue").text if media_node.find("DisplayValue") is not None else ""
        else:
            media = ""

        # BusProtocol (6 → SAS)
        proto_node = inst.find(".//PROPERTY[@NAME='BusProtocol']")
        if proto_node is not None:
            proto_value = proto_node.find("VALUE").text if proto_node.find("VALUE") is not None else ""
            if proto_value == "6":
                proto = "SAS"
            else:
                proto = proto_node.find("DisplayValue").text if proto_node.find("DisplayValue") is not None else ""
        else:
            proto = ""

        # MaxCapableSpeed (цифра → DisplayValue)
        speed_node = inst.find(".//PROPERTY[@NAME='MaxCapableSpeed']")
        if speed_node is not None:
            speed_value = speed_node.find("VALUE").text if speed_node.find("VALUE") is not None else ""
            if speed_value.isdigit() and len(speed_value) == 1:
                speed = speed_node.find("DisplayValue").text if speed_node.find("DisplayValue") is not None else ""
            else:
                speed = speed_value
        else:
            speed = ""

        # DriveFormFactor (2 → 2.5 inch)
        form_node = inst.find(".//PROPERTY[@NAME='DriveFormFactor']")
        if form_node is not None:
            form_value = form_node.find("VALUE").text if form_node.find("VALUE") is not None else ""
            if form_value == "2":
                form = "2.5 inch"
            else:
                form = form_node.find("DisplayValue").text if form_node.find("DisplayValue") is not None else ""
        else:
            form = ""

        description = " ".join(filter(None, [size_gb, media, proto, speed, form]))
        key = (description, part_number)
        counter[key] = counter.get(key, 0) + 1

    return [(descr, spare, qty) for (descr, spare), qty in sorted(counter.items())]

def _collect_pci_devices(root: etree._Element) -> List[Tuple[str, str, int]]:
    """
    Собирает информацию о PCI-устройствах, исключая встроенные (Embedded/Integrated).
    Возвращает список кортежей: (Description, "PCI Device", 1)
    """
    devices = []
    for inst in root.findall(".//INSTANCE[@CLASSNAME='DCIM_PCIDeviceView']"):
        # Проверяем InstanceID на наличие ключевых слов
        instance_id_node = inst.find(".//PROPERTY[@NAME='InstanceID']/VALUE")
        if instance_id_node is not None and instance_id_node.text:
            instance_id = instance_id_node.text.strip()
            if "Embedded" in instance_id or "Integrated" in instance_id:
                continue  # Пропускаем встроенные устройства

            if "NIC" in instance_id:
                for inst2 in root.findall(".//INSTANCE[@CLASSNAME='DCIM_NICView']"):
                    part_num = inst2.find(".//PROPERTY[@NAME='PartNumber']/VALUE")
                    desc_node = inst2.find(".//PROPERTY[@NAME='DeviceDescription']/VALUE")
                    if desc_node is not None and desc_node.text:
                        desc = desc_node.text.strip()
                        devices.append((desc, part_num.text.strip() if part_num is not None else "", 1))
                continue  # NIC уже обработаны, пропускаем их как PCI устройства

            if "FC" in instance_id:
                for inst2 in root.findall(".//INSTANCE[@CLASSNAME='DCIM_FCView']"):
                    part_num = inst2.find(".//PROPERTY[@NAME='PartNumber']/VALUE")
                    desc_node = inst2.find(".//PROPERTY[@NAME='DeviceDescription']/VALUE")
                    if desc_node is not None and desc_node.text:
                        desc = desc_node.text.strip()
                        devices.append((desc, part_num.text.strip() if part_num is not None else "", 1))
                continue  # FC адаптеры уже обработаны, пропускаем их как PCI устройства

        # Если устройство не встроенное - добавляем его описание
        desc_node = inst.find(".//PROPERTY[@NAME='Description']/VALUE")
        if desc_node is not None and desc_node.text:
            desc = desc_node.text.strip()
            devices.append((desc, "part_num", 1))

    return devices

def parse_xml(file_path: Path) -> Tuple[
    Optional[str],
    Optional[str],
    Optional[str],  # Добавляем парт-номер материнской платы
    List[Tuple[str, str, int]],
]:
    """Возвращает SN, модель, парт-номер мат. платы и список всех найденных компонентов."""
    tree = etree.parse(str(file_path))
    root = tree.getroot()

    sn: Optional[str] = _get_sn_from_system_view(root)
    model: Optional[str] = _get_model_from_system_view(root)
    motherboard_pn: Optional[str] = _get_motherboard_part_number(root)  # Новое поле

    components: List[Tuple[str, str, int]] = []
    components.extend(_collect_memory(root))
    components.extend(_collect_cpus(root))
    components.extend(_collect_psus(root))
    components.extend(_collect_disks(root))
    components.extend(_collect_pci_devices(root))

    return sn, model, motherboard_pn, components


def write_excel(
        out_path: Path,
        data: List[Tuple[Optional[str], Optional[str], Optional[str], List[Tuple[str, str, int]]]],
) -> None:
    """Записывает собранные данные в Excel-файл с оформлением."""
    from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    wb = Workbook()
    ws = wb.active
    ws.title = "Servers"

    # Стили границ и оформления
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'),
                          top=Side(style='thick'), bottom=Side(style='thick'))
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    summary_fill = PatternFill("solid", fgColor="EEEEEE")

    # Заголовки
    headers = ["SN", "Model", "Motherboard PN", "Description", "Spare", "Quant"]
    ws.append(headers)

    # Стиль для заголовков
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.fill = header_fill

    # Собираем статистику по запчастям
    spare_stats = defaultdict(int)
    for _, _, _, components in data:
        for _, spare, qty in components:
            if spare and spare != "PCI Device":  # Исключаем PCI устройства из подсчета
                spare_stats[spare] += qty

    # Заполнение данных
    row = 2
    for sn, model, motherboard_pn, components in data:
        if not components:
            continue

        start_row = row
        end_row = row + len(components) - 1

        # Записываем данные для каждого компонента
        for descr, spare, qty in components:
            # Для PCI устройств меняем местами Description и Spare
            if spare == "PCI Device":
                ws.append([sn, model, motherboard_pn, spare, descr, qty])  # Меняем местами
            else:
                ws.append([sn, model, motherboard_pn, descr, spare, qty])  # Обычный порядок
            row += 1

        # Объединяем ячейки SN, Model и Motherboard PN для этого сервера
        for col in [1, 2, 3]:  # Колонки A (SN), B (Model), C (Motherboard PN)
            ws.merge_cells(start_row=start_row, start_column=col,
                           end_row=end_row, end_column=col)

            # Центрируем объединённые ячейки
            merged_cell = ws.cell(row=start_row, column=col)
            merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Применяем стили ко всем ячейкам
        for r in range(start_row, end_row + 1):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

        # Жирные границы вокруг блока сервера
        for r in range(start_row, end_row + 1):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=col)

                # Границы блока
                if r == start_row:
                    new_border = copy(cell.border)
                    new_border.top = Side(style='thick')
                    cell.border = new_border
                if r == end_row:
                    new_border = copy(cell.border)
                    new_border.bottom = Side(style='thick')
                    cell.border = new_border
                if col == 1:
                    new_border = copy(cell.border)
                    new_border.left = Side(style='thick')
                    cell.border = new_border
                if col == len(headers):
                    new_border = copy(cell.border)
                    new_border.right = Side(style='thick')
                    cell.border = new_border

    # Добавляем сводку по запчастям
    if spare_stats:
        summary_col = len(headers) + 2
        summary_row = 2

        # Заголовок сводки
        ws.cell(row=summary_row, column=summary_col, value="Сводка по запчастям").font = Font(bold=True)
        summary_row += 1

        # Сортируем запчасти по количеству (по убыванию)
        sorted_spares = sorted(spare_stats.items(), key=lambda x: x[1], reverse=True)

        # Добавляем данные
        for spare, total in sorted_spares:
            ws.cell(row=summary_row, column=summary_col, value=spare)
            ws.cell(row=summary_row, column=summary_col + 1, value=total).alignment = Alignment(horizontal='right')
            summary_row += 1

        # Стили для сводки
        for r in range(2, summary_row):
            ws.cell(row=r, column=summary_col).border = thin_border
            ws.cell(row=r, column=summary_col + 1).border = thin_border
            if r == 2:
                ws.cell(row=r, column=summary_col).fill = header_fill
                ws.cell(row=r, column=summary_col).font = Font(bold=True)
            else:
                ws.cell(row=r, column=summary_col).fill = summary_fill

        # Объединяем заголовок
        ws.merge_cells(start_row=2, start_column=summary_col,
                       end_row=2, end_column=summary_col + 1)

    # Автоподбор ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Фиксированные минимальные ширины
    ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width, 20)  # SN
    ws.column_dimensions['B'].width = max(ws.column_dimensions['B'].width, 25)  # Model
    ws.column_dimensions['C'].width = max(ws.column_dimensions['C'].width, 20)  # Motherboard PN
    ws.column_dimensions['D'].width = max(ws.column_dimensions['D'].width, 45)  # Description/Spare
    ws.column_dimensions['E'].width = max(ws.column_dimensions['E'].width, 25)  # Spare/Description

    wb.save(out_path)

def main() -> None:
    base_dir = ask_path_or_arg()
    out_dir = base_dir / OUT_SUBDIR
    out_dir.mkdir(parents=True, exist_ok=True)

    # 1️⃣ Распаковываем все XML-файлы
    zip_files = [p for p in base_dir.iterdir() if p.is_file() and p.suffix.lower() == ".zip"]
    if not zip_files:
        LOGGER.warning("No zip archives found in %s", base_dir)

    for zip_path in zip_files:
        extract_from_archive(zip_path, out_dir)

    # 2️⃣ Парсим каждый найденный XML-файл
    parsed: List[Tuple[Optional[str], Optional[str], Optional[str], List[Tuple[str, str, int]]]] = []
    for xml_path in out_dir.glob("*" + TARGET_NAME):
        LOGGER.info("Parsing %s", xml_path.name)
        try:
            sn, model, motherboard_pn, components = parse_xml(xml_path)
        except (etree.XMLSyntaxError, OSError, ValueError) as exc:
            LOGGER.error("Failed to parse %s: %s", xml_path.name, exc)
            continue
        parsed.append((sn or "", model or "", motherboard_pn or "", components))

    # 3️⃣ Записываем результат в Excel-файл
    out_excel = out_dir / "servers_report.xlsx"
    write_excel(out_excel, parsed)


if __name__ == "__main__":
    main()