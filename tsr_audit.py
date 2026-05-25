import io
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path
import pandas as pd
import re

TSR_DIR = Path('./tsr')
archives = sorted(TSR_DIR.glob('*.zip'))
print(f'Найдено {len(archives)} архивов TSR в {TSR_DIR.resolve()}')
for archive in archives[:10]:
    print('-', archive.name)


VIEW_PATH = 'tsr/hardware/sysinfo/inventory/sysinfo_DCIM_View.xml'
COMPONENT_CLASSES = {
    'DCIM_CPUView': 'CPU',
    'DCIM_MemoryView': 'DIMM',
    'DCIM_PowerSupplyView': 'PS',
    'DCIM_PhysicalDiskView': 'Drive',
    'DCIM_PCIDeviceView': 'PCI',
}

def get_value(prop):
    if prop is None:
        return ''
    value = prop.find('VALUE')
    if value is not None and value.text is not None:
        return value.text.strip()
    display = prop.find('DisplayValue')
    return display.text.strip() if display is not None and display.text else ''


def instance_props(inst):
    props = {}
    for prop in inst.findall('PROPERTY'):
        name = prop.attrib.get('NAME')
        if not name:
            continue
        props[name] = get_value(prop)
    return props


def open_nested_tsr(outer_zip_path):
    with zipfile.ZipFile(outer_zip_path, 'r') as outer:
        nested = [name for name in outer.namelist() if name.lower().endswith('.pl.zip')]
        if not nested:
            raise FileNotFoundError(f'No nested .pl.zip found in {outer_zip_path}')
        with outer.open(nested[0]) as nested_file:
            return zipfile.ZipFile(io.BytesIO(nested_file.read()), 'r')


def parse_tsr_archive(outer_zip_path):
    with open_nested_tsr(outer_zip_path) as nested:
        if VIEW_PATH not in nested.namelist():
            raise FileNotFoundError(f'{VIEW_PATH} not found inside {outer_zip_path}')
        text = nested.read(VIEW_PATH)
    root = ET.fromstring(text)

    system_instances = [inst for inst in root.findall('.//INSTANCE') if inst.attrib.get('CLASSNAME') == 'DCIM_SystemView']
    server = instance_props(system_instances[0]) if system_instances else {}
    server_model = server.get('Model', 'Unknown')
    server_serial = server.get('ServiceTag', server.get('ChassisServiceTag', 'Unknown'))
    server_part = server.get('BoardPartNumber', server.get('PartNumber', 'Unknown'))

    components = defaultdict(list)
    for inst in root.findall('.//INSTANCE'):
        cname = inst.attrib.get('CLASSNAME')
        if cname in COMPONENT_CLASSES:
            components[cname].append(instance_props(inst))

    return server_model, server_serial, server_part, components


#   group_key - поля, по которым объединяются записи в одну строку
#   part_fields - поля, которые используются как партийный номер компонента
#   desc_fields - поля, которые формируют описание компонента
COMPONENT_ROW_CONFIG = {
    'CPU': {
        'group_key': ['PPID', 'PartNumber', 'ProcessorID', 'Model', 'Description'],
        'part_fields': ['PPID', 'PartNumber', 'ProcessorID', 'FQDD', 'SerialNumber'],
        'desc_fields': ['Model', 'SocketDesignation', 'ProcessorType'],
    },
    'DIMM': {
        'group_key': ['PPID', 'PartNumber', 'DeviceLocator', 'SerialNumber'],
        'part_fields': ['PPID', 'PartNumber', 'SerialNumber', 'FQDD'],
        'desc_fields': ['DeviceLocator', 'Model', 'Description', 'PartNumber', 'Capacity'],
    },
    'PS': {
        'group_key': ['PPID', 'PartNumber', 'FQDD', 'DeviceDescription'],
        'part_fields': ['PPID', 'PartNumber', 'FQDD'],
        'desc_fields': ['Model', 'Description'],
    },
    'Drive': {
        'group_key': ['PPID', 'PartNumber', 'FQDD', 'Model', 'DeviceDescription'],
        'part_fields': ['Part Number','PPID', 'PartNumber', 'SerialNumber', 'FQDD'],
        'desc_fields': ['Model'],
    },
    'PCI': {
        'group_key': ['PPID', 'Model', 'Description'],
        'part_fields': ['PartNumber', 'PPID', 'FQDD', 'SerialNumber'],
        'desc_fields': ['Description', 'Model'],
    },
}


def build_component_rows(server_model, server_serial, server_part, components):
    rows = []

    def get_config(comp_type):
        return COMPONENT_ROW_CONFIG.get(comp_type, {
            'group_key': ['PPID', 'PartNumber', 'Part Number', 'Model', 'Description'],
            'part_fields': ['PPID', 'PartNumber', 'Part Number', 'SerialNumber', 'FQDD'],
            'desc_fields': ['Description', 'DeviceDescription', 'Model'],
        })

    def should_skip(item, comp_type):
        if comp_type != 'PCI':
            return False
        text = ' '.join(
            str(item.get(field, '')).lower() for field in ['DeviceDescription', 'Description', 'Model', 'FQDD']
        )
        return any(word in text for word in ['embedded', 'integrated'])

    def group_items(cname):
        comp_type = COMPONENT_CLASSES[cname]
        config = get_config(comp_type)
        grouped = defaultdict(list)
        for item in components.get(cname, []):
            if should_skip(item, comp_type):
                continue
            key = next((item.get(field) for field in config['group_key'] if item.get(field)), comp_type)
            grouped[key].append(item)
        for key, items in grouped.items():
            sample = items[0]
            part = next((sample.get(field) for field in config['part_fields'] if sample.get(field)), 'N/A')
            descriptions = [sample.get(field) for field in config['desc_fields'] if sample.get(field)]
            desc = '; '.join(descriptions) or key or comp_type
            rows.append({
                'component': comp_type,
                'quantity': len(items),
                'component_part': part,
                'description': desc,
            })

    for cname in COMPONENT_CLASSES:
        group_items(cname)
    return rows


def extract_drive_part_number(ppid_str):
    if not ppid_str or pd.isna(ppid_str):
        return 'N/A'
    
    ppid_str = str(ppid_str).strip().upper()

    match_with_dash = re.match(r'^[A-Z]{2}-([A-Z0-9]{5,6})-', ppid_str)
    if match_with_dash:
        return match_with_dash.group(1)
        
    match_no_dash = re.match(r'^[A-Z]{2}([A-Z0-9]{5,6})[A-Z0-9]', ppid_str)
    if match_no_dash:
        return match_no_dash.group(1)
        
    return 'N/A'


table_rows = []
server_index = 0
for archive in archives:
    try:
        server_model, server_serial, server_part, components = parse_tsr_archive(archive)
    except Exception as exc:
        print(f'Failed parse {archive.name}: {exc}')
        continue
    server_index += 1
    component_rows = build_component_rows(server_model, server_serial, server_part, components)
    first = True
    for comp_row in component_rows:
        # Извлекаем партийник компонента
        if comp_row['component'] == "Drive":
            part_number = extract_drive_part_number(comp_row['component_part'])
        else:
            part_number = comp_row['component_part']
        
        table_rows.append({
            'номер': server_index if first else '',
            'название сервера': server_model if first else '',
            'серийный номер сервера': server_serial if first else '',
            'партийный номер сервера': server_part[1:-3] if first else '',
            'компонент': comp_row['component'],
            'количество': comp_row['quantity'],
            'партийный номер компонента': part_number,
            'описание компонента': comp_row['description'],
        })

        first = False

df = pd.DataFrame(table_rows)


# диски до и после merge
print("Диски до merge:")
drive_rows_before = [row for row in table_rows if row['компонент'] == 'Drive']
print(f"Всего строк с дисками: {len(drive_rows_before)}")
for row in drive_rows_before:
    print(f"  {row['номер']} | {row['партийный номер компонента']} | кол-во: {row['количество']}")

print("\nДиски после merge:")
drive_rows_after = [row for row in table_rows if row['компонент'] == 'Drive']
print(f"Всего строк с дисками: {len(drive_rows_after)}")
for row in drive_rows_after:
    print(f"  {row['номер']} | {row['партийный номер компонента']} | кол-во: {row['количество']}")

from openpyxl import load_workbook

output_dir = Path('xl')
output_dir.mkdir(parents=True, exist_ok=True)
output_csv = output_dir / 'dell_tsr_audit.csv'
output_excel = output_dir / 'dell_tsr_audit.xlsx'
df.to_csv(output_csv, index=False, encoding='utf-8-sig')
try:
    df.to_excel(output_excel, index=False)

    wb = load_workbook(output_excel)
    ws = wb.active
    if len(df) > 0:
        merge_columns = ['A', 'B', 'C', 'D']
        current_start = 2
        for idx in range(1, len(df)):
            if df.loc[idx, 'номер'] != '':
                if current_start < idx + 1:
                    for col in merge_columns:
                        ws.merge_cells(f'{col}{current_start}:{col}{idx + 1}')
                current_start = idx + 2
        last_row = len(df) + 1
        if current_start < last_row:
            for col in merge_columns:
                ws.merge_cells(f'{col}{current_start}:{col}{last_row}')
    wb.save(output_excel)
except Exception as exc:
    print('Excel export failed:', exc)
print('Saved audit table to', output_csv)
print('Saved audit table to', output_excel)